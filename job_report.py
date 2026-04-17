#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║         Job Seeker Prompt Generator                     ║
║  Excel + Resume → Ready-to-paste AI prompts             ║
╚══════════════════════════════════════════════════════════╝

Reads your Excel file and resume files, injects your real data into
each prompt template, then:
  - Prints every prompt to the terminal
  - Saves a clean .md file per company you can open and paste from

No API key required. Paste each prompt directly into ChatGPT, Claude,
Gemini, or any other web AI interface.

Run `python job_report.py --save-config` to generate a starter config.json.

Usage:
  python job_report.py [OPTIONS]

Options:
  --excel          PATH   Path to Excel file        (default: from config or jobs_data.xlsx)
  --resume         PATH   Path to resume .txt file  (default: from config)
  --resume-summary PATH   Path to summary .txt file (default: from config)
  --out-dir        PATH   Directory for .md files   (default: from config or ./prompts)
  --config         PATH   Path to config file       (default: ./config.json)
  --save-config           Write a starter config.json and exit
  --print-only            Print prompts to terminal only, do not save files
  --help                  Show this help message

Excel Schema (one row per job application):
  Column A - Company         Company name
  Column B - Job_Description Full job description text

Examples:
  # First-time setup
  python job_report.py --save-config

  # Generate prompts for all rows in the spreadsheet
  python job_report.py

  # Just print to screen without saving files
  python job_report.py --print-only

  # Use a different resume for a one-off run
  python job_report.py --resume ~/docs/resume_v2.txt
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


# ─────────────────────────── CONFIG ────────────────────────────────────────

CONFIG_DEFAULTS = {
    "excel":          "jobs_data.xlsx",
    "resume":         "",
    "resume_summary": "",
    "out_dir":        "prompts",
}

CONFIG_HEADER = (
    "// Job Seeker Prompt Generator - Configuration\n"
    "// All paths may be absolute or relative to this file.\n"
    "// CLI flags always override values set here.\n"
    "//\n"
    "// Fields:\n"
    "//   excel          - Excel file with Company + Job_Description columns\n"
    "//   resume         - Your full resume as a plain .txt file\n"
    "//   resume_summary - Your 2-3 sentence professional summary as a .txt file\n"
    "//   out_dir        - Folder where generated prompt .md files are saved\n"
)


def load_config(config_path: Path) -> dict:
    """Load config.json, stripping // comment lines before parsing."""
    if not config_path.exists():
        return {}
    raw = config_path.read_text(encoding="utf-8")
    stripped = "\n".join(
        line for line in raw.splitlines()
        if not line.strip().startswith("//")
    )
    try:
        return json.loads(stripped)
    except json.JSONDecodeError as e:
        sys.exit(f"Could not parse {config_path}: {e}")


def save_starter_config(config_path: Path) -> None:
    if config_path.exists():
        answer = input(f"  {config_path} already exists. Overwrite? [y/N] ").strip().lower()
        if answer != "y":
            print("Aborted.")
            return
    data = json.dumps(CONFIG_DEFAULTS, indent=2)
    config_path.write_text(CONFIG_HEADER + data + "\n", encoding="utf-8")
    print(f"Starter config written -> {config_path.resolve()}")
    print("Edit it to set your resume paths, then run: python job_report.py")


def resolve_setting(cli_val, config: dict, key: str, fallback=None):
    """Return CLI value if set, else config value, else fallback."""
    if cli_val is not None and cli_val != "":
        return cli_val
    return config.get(key) or fallback


# ─────────────────────────── RUBRIC ────────────────────────────────────────

RUBRIC = [
    {
        "id": 1,
        "rule": "Specificity",
        "description": (
            "The response must reference specific details from the job description, "
            "company name, or resume - not generic advice."
        ),
    },
    {
        "id": 2,
        "rule": "Actionability",
        "description": (
            "Every answer must include at least one concrete, actionable takeaway "
            "the candidate can act on immediately."
        ),
    },
    {
        "id": 3,
        "rule": "Completeness",
        "description": (
            "The response must fully address the question asked without leaving "
            "major sub-points unanswered."
        ),
    },
    {
        "id": 4,
        "rule": "Candidate-Centric Framing",
        "description": (
            "Insights must be framed from the job seeker's perspective - helping "
            "them prepare, position themselves, or improve."
        ),
    },
    {
        "id": 5,
        "rule": "Professional Tone",
        "description": (
            "Language must be clear, professional, and free of filler phrases like "
            "'Great question!' or vague platitudes."
        ),
    },
]


# ─────────────────────────── PROMPTS ───────────────────────────────────────

def build_prompts(
    company: str,
    job_desc: str,
    resume_summary: str,
    resume: str,
) -> list:
    """Build the five prompts with real data injected."""
    return [
        {
            "id":    "culture",
            "label": "1. Company Culture & Intel",
            "instructions": (
                "Paste this into your AI of choice. It will summarize the company's "
                "culture, known challenges, and recent news to help you prepare."
            ),
            "prompt": (
                f"Summarize {company}'s culture, challenges, and recent news "
                "for a job interview. Be specific about what a candidate should "
                "know walking into an interview. Include cultural values, known "
                "workplace challenges, recent news or leadership changes, and one "
                "smart question the candidate could ask that shows insider knowledge."
            ),
        },
        {
            "id":    "hidden_requirements",
            "label": "2. Hidden Job Requirements",
            "instructions": (
                "Paste this prompt to uncover the unstated priorities and personality "
                "traits the hiring manager actually wants."
            ),
            "prompt": (
                "What is this job actually looking for underneath the formal "
                "requirements? Read between the lines and surface the unstated "
                "priorities, pain points, and personality traits the hiring manager "
                "really wants.\n\n"
                f"--- JOB DESCRIPTION ---\n{job_desc}"
            ),
        },
        {
            "id":    "gap_analysis",
            "label": "3. Gap Analysis",
            "instructions": (
                "Paste this prompt to get an honest assessment of where your "
                "background is weakest for this role, with reframing suggestions."
            ),
            "prompt": (
                "Here is a job description and my background. Where am I weakest "
                "for this role? Be honest and specific. For each gap, suggest a "
                "concrete way to address or reframe it.\n\n"
                f"--- JOB DESCRIPTION ---\n{job_desc}\n\n"
                f"--- MY RESUME ---\n{resume}"
            ),
        },
        {
            "id":    "resume_summary",
            "label": "4. Tailored Resume Summary",
            "instructions": (
                "Paste this prompt to get a rewritten resume summary that mirrors "
                "the language and priorities of the job description."
            ),
            "prompt": (
                f"Rewrite my resume summary specifically for the role at {company}. "
                "Mirror the language and priorities in the job description. "
                "Keep it to 3-4 sentences, first-person, no fluff.\n\n"
                f"--- JOB DESCRIPTION ---\n{job_desc}\n\n"
                f"--- MY CURRENT SUMMARY ---\n{resume_summary}"
            ),
        },
        {
            "id":    "rubric_eval",
            "label": "5. Rubric Evaluation (paste AFTER getting a response)",
            "instructions": (
                "After the AI answers any of the above prompts, paste this as a "
                "follow-up to score the response against 5 quality rules."
            ),
            "prompt": (
                "Please evaluate your previous response against this rubric. "
                "For each rule give PASS or FAIL with one sentence of reasoning. "
                "End with an overall score out of 5 and one improvement suggestion.\n\n"
                "RUBRIC:\n"
                + "\n".join(
                    f"{r['id']}. {r['rule']}: {r['description']}"
                    for r in RUBRIC
                )
                + "\n\n"
                "Format:\n"
                "1. Specificity: [PASS/FAIL] - <reasoning>\n"
                "2. Actionability: [PASS/FAIL] - <reasoning>\n"
                "3. Completeness: [PASS/FAIL] - <reasoning>\n"
                "4. Candidate-Centric Framing: [PASS/FAIL] - <reasoning>\n"
                "5. Professional Tone: [PASS/FAIL] - <reasoning>\n\n"
                "Overall Score: X/5\n"
                "Improvement Suggestion: <one sentence>"
            ),
        },
    ]


# ─────────────────────────── FILE LOADERS ──────────────────────────────────

def load_text_file(path: str, label: str) -> str:
    p = Path(path).expanduser().resolve()
    if not p.exists():
        sys.exit(f"ERROR: {label} file not found: {p}")
    if not p.is_file():
        sys.exit(f"ERROR: {label} path is not a file: {p}")
    content = p.read_text(encoding="utf-8").strip()
    if not content:
        sys.exit(f"ERROR: {label} file is empty: {p}")
    return content


# ─────────────────────────── EXCEL READER ──────────────────────────────────

def read_excel(path: str) -> list:
    p = Path(path).expanduser().resolve()
    if not p.exists():
        sys.exit(f"Excel file not found: {p}")

    wb = openpyxl.load_workbook(p)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    required = {"Company", "Job_Description"}
    missing = required - set(headers)
    if missing:
        sys.exit(
            f"Excel is missing required columns: {missing}\n"
            f"Found: {headers}"
        )

    col_map = {h: i for i, h in enumerate(headers)}
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        company = row[col_map["Company"]]
        if not company:
            continue
        rows.append({
            "row":             row_idx,
            "company":         str(company).strip(),
            "job_description": str(row[col_map["Job_Description"]] or "").strip(),
        })

    if not rows:
        sys.exit("No data rows found in the Excel file.")
    return rows


# ─────────────────────────── OUTPUT BUILDERS ───────────────────────────────

DIVIDER      = "=" * 72
THIN_DIVIDER = "-" * 72
PROMPT_FENCE = "~" * 72   # visually distinct so the copy boundary is obvious


def print_company_prompts(company: str, prompts: list) -> None:
    """Print all prompts for one company to stdout."""
    print(f"\n{DIVIDER}")
    print(f"  PROMPTS FOR: {company.upper()}")
    print(DIVIDER)

    for item in prompts:
        print(f"\n{THIN_DIVIDER}")
        print(f"  {item['label']}")
        print(f"  {item['instructions']}")
        print(THIN_DIVIDER)
        print()
        # Fence the prompt so it is visually distinct from surrounding text
        print(PROMPT_FENCE)
        print(item["prompt"])
        print(PROMPT_FENCE)
        print()

    print(f"{DIVIDER}\n")


def build_markdown(
    company: str,
    prompts: list,
    timestamp: str,
    resume_path: str,
    summary_path: str,
) -> str:
    """Build a clean markdown document with every prompt ready to copy."""

    lines = [
        f"# Job Application Prompts — {company}",
        f"",
        f"*Generated: {timestamp}*  ",
        f"*Resume: `{resume_path}`*  ",
        f"*Summary: `{summary_path}`*",
        f"",
        f"---",
        f"",
        f"## How to use this file",
        f"",
        f"Each section below contains one ready-to-paste prompt.",
        f"Copy the text inside the code block and paste it into",
        f"[ChatGPT](https://chat.openai.com), [Claude](https://claude.ai),",
        f"[Gemini](https://gemini.google.com), or any other AI web interface.",
        f"Run the prompts in order for best results.",
        f"",
        f"---",
        f"",
        f"## Quality Rubric",
        f"",
        f"Every response you receive should meet these 5 rules.",
        f"Use **Prompt 5** to ask the AI to self-evaluate.",
        f"",
        f"| # | Rule | What it checks |",
        f"|---|------|----------------|",
    ]

    for r in RUBRIC:
        lines.append(f"| {r['id']} | **{r['rule']}** | {r['description']} |")

    lines += ["", "---", ""]

    for item in prompts:
        lines += [
            f"## {item['label']}",
            f"",
            f"> {item['instructions']}",
            f"",
            f"```",
            item["prompt"],
            f"```",
            f"",
            f"---",
            f"",
        ]

    return "\n".join(lines)


# ─────────────────────────── CLI & MAIN ────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Job Seeker Prompt Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--excel",          default=None,
                        help="Path to Excel file")
    parser.add_argument("--resume",         default=None,
                        help="Path to resume .txt file")
    parser.add_argument("--resume-summary", default=None, dest="resume_summary",
                        help="Path to resume summary .txt file")
    parser.add_argument("--out-dir",        default=None, dest="out_dir",
                        help="Output directory for prompt files")
    parser.add_argument("--config",         default="config.json",
                        help="Path to config file (default: ./config.json)")
    parser.add_argument("--save-config",    action="store_true",
                        help="Write a starter config.json and exit")
    parser.add_argument("--print-only",     action="store_true", dest="print_only",
                        help="Print prompts to terminal only, do not save files")
    return parser.parse_args()


def main() -> None:
    args        = parse_args()
    config_path = Path(args.config).expanduser().resolve()

    if args.save_config:
        save_starter_config(config_path)
        return

    config       = load_config(config_path)
    excel_path   = resolve_setting(args.excel,          config, "excel",          CONFIG_DEFAULTS["excel"])
    resume_path  = resolve_setting(args.resume,         config, "resume",         "")
    summary_path = resolve_setting(args.resume_summary, config, "resume_summary", "")
    out_dir_str  = resolve_setting(args.out_dir,        config, "out_dir",        CONFIG_DEFAULTS["out_dir"])

    # Validate resume paths
    errors = []
    if not resume_path:
        errors.append('  --resume PATH  or  set "resume" in config.json')
    if not summary_path:
        errors.append('  --resume-summary PATH  or  set "resume_summary" in config.json')
    if errors:
        sys.exit(
            "Resume file paths not provided. Supply them via:\n"
            + "\n".join(errors)
            + "\n\nTip: run  python job_report.py --save-config  to create a starter config.json"
        )

    resume_text  = load_text_file(resume_path,  "Resume")
    summary_text = load_text_file(summary_path, "Resume Summary")

    if not args.print_only:
        out_dir = Path(out_dir_str).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)

    cfg_label = str(config_path) if config_path.exists() else "(not found - using defaults)"

    print(f"\n{'─'*64}")
    print(f"  Job Seeker Prompt Generator")
    print(f"  Excel          : {excel_path}")
    print(f"  Resume         : {resume_path}")
    print(f"  Resume Summary : {summary_path}")
    if not args.print_only:
        print(f"  Output folder  : {out_dir_str}")
    print(f"  Config         : {cfg_label}")
    print(f"{'─'*64}")

    rows      = read_excel(excel_path)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\n  Found {len(rows)} job row(s).\n")

    for row in rows:
        company = row["company"]
        prompts = build_prompts(
            company        = company,
            job_desc       = row["job_description"],
            resume_summary = summary_text,
            resume         = resume_text,
        )

        # Always print to screen
        print_company_prompts(company, prompts)

        # Optionally save .md file
        if not args.print_only:
            safe_company = "".join(
                c if c.isalnum() or c in "-_" else "_" for c in company
            )
            date_slug = datetime.now().strftime("%Y%m%d_%H%M")
            out_file  = out_dir / f"{safe_company}_{date_slug}.md"

            md_content = build_markdown(
                company      = company,
                prompts      = prompts,
                timestamp    = timestamp,
                resume_path  = str(Path(resume_path).resolve()),
                summary_path = str(Path(summary_path).resolve()),
            )
            out_file.write_text(md_content, encoding="utf-8")
            print(f"  Saved -> {out_file.resolve()}\n")

    print(f"{'─'*64}")
    if args.print_only:
        print("  Done. (print-only mode — no files saved)")
    else:
        print(f"  Done. Prompt files saved to: {out_dir.resolve()}")
    print(f"{'─'*64}\n")


if __name__ == "__main__":
    main()
